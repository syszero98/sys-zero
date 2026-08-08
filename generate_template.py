#!/usr/bin/env python3
"""
Generate Sample Excel Template
Run this script to create a sample Excel file for data import
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_sample_template():
    """Create a sample Excel template for student results"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Results"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Add headers
    headers = ['Roll', 'Name', 'Bangla', 'English', 'Math', 'Science']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add sample data
    sample_data = [
        [1, 'আহমেদ হোসেন', 85, 78, 92, 88],
        [2, 'ফাতিমা খান', 90, 88, 95, 92],
        [3, 'করিম সাহেব', 75, 82, 80, 78],
        [4, 'রহিমা বেগম', 88, 85, 90, 87],
        [5, 'হাসান মিয়া', 78, 76, 82, 80],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            # Format numbers as numbers
            if col_idx > 2:  # Subject columns
                cell.number_format = '0'
    
    # Add instruction row
    instruction_row = len(sample_data) + 3
    instruction = ws.cell(row=instruction_row, column=1)
    instruction.value = "⚠️ Instructions: Keep headers same. Add more rows as needed. Marks should be 0-100."
    instruction.font = Font(italic=True, color="FF0000", size=10)
    
    # Save template
    template_path = os.path.join(
        os.path.dirname(__file__),
        'sample_template.xlsx'
    )
    
    wb.save(template_path)
    print(f"✅ Sample template created: {template_path}")
    return template_path

if __name__ == '__main__':
    create_sample_template()
