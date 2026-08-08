"""
Excel Parser Utility
Handles Excel file parsing and data extraction
"""

import os
from openpyxl import load_workbook
import json

def parse_excel_file(file_path):
    """
    Parse Excel file and extract student data
    Expected columns: Roll, Name, Subject1, Subject2, ...
    """
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        # Extract headers
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers or 'Roll' not in headers and 'রোল' not in headers:
            return None, "Invalid Excel format. Must contain 'Roll' column."
        
        # Find Roll and Name columns
        roll_col = None
        name_col = None
        
        for i, header in enumerate(headers):
            if header.lower() in ['roll', 'রোল']:
                roll_col = i
            elif header.lower() in ['name', 'নাম']:
                name_col = i
        
        if roll_col is None:
            return None, "Roll column not found"
        
        if name_col is None:
            return None, "Name column not found"
        
        # Extract student data
        students = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[roll_col]:  # Skip empty rows
                continue
            
            student = {
                'roll': str(row[roll_col]).strip(),
                'name': str(row[name_col]).strip() if name_col < len(row) else '',
                'subjects': []
            }
            
            # Extract subject marks
            for col_idx, header in enumerate(headers):
                if col_idx not in [roll_col, name_col] and header.lower() not in ['roll', 'রোল', 'name', 'নাম']:
                    try:
                        marks = float(row[col_idx]) if col_idx < len(row) and row[col_idx] else 0
                        grade = get_grade(marks)
                        
                        student['subjects'].append({
                            'name': header,
                            'marks': marks,
                            'grade': grade
                        })
                    except:
                        pass
            
            students.append(student)
        
        return students, "Success"
    
    except Exception as e:
        return None, f"Error parsing Excel: {str(e)}"

def get_grade(marks):
    """Calculate grade based on marks"""
    if marks >= 80:
        return 'A'
    elif marks >= 70:
        return 'B'
    elif marks >= 60:
        return 'C'
    elif marks >= 50:
        return 'D'
    elif marks >= 40:
        return 'E'
    else:
        return 'F'

def validate_file(file_path):
    """Validate if file is valid Excel"""
    if not os.path.exists(file_path):
        return False, "File not found"
    
    if not file_path.endswith(('.xlsx', '.xls')):
        return False, "Invalid file format. Use .xlsx or .xls"
    
    return True, "Valid"
